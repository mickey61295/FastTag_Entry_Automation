from openpyxl.styles import Color, PatternFill, Font, Border
import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl, time, datetime
from selenium import webdriver
from selenium.webdriver.common.keys import Keys

root = tk.Tk()
root.withdraw()
filename = filedialog.askopenfilename(initialdir="/",title="Open File",filetypes=(("Excel", "*.xlsx"), ("All Files", "*.*")))
wb = openpyxl.load_workbook(filename)
sh = wb['Sheet1']
tran = []
stat = []
amount = []
count = sh.max_row+1
for i in range(2, count):
    tran.append(sh.cell(row=i, column=1).value)

chromedriver = filedialog.askopenfilename(initialdir="/",title="Open File",filetypes=(("Executable", "*.exe"), ("All Files", "*.*")))
browser = webdriver.Chrome(chromedriver)
browser.get('https://fastagcsc.bankofbaroda.com/BOBPOS/Default.aspx')
user_n = browser.find_element_by_id('txtUserName')
user_n_in = input('Enter Your User id')
user_n.send_keys(user_n_in)
pass_in = input('Enter your Password')
pass_w = browser.find_element_by_id(pass_in)
pass_w.send_keys('Welcome@2716')
result = messagebox.askokcancel('Captcha', 'Navigate to browser window and Enter Captcha.\nPress Ok once Done.')
browser.get('https://fastagcsc.bankofbaroda.com/BOBPOS/pages/Retailer/MakePayment/MPOSDeposit.aspx')
for i in range(0,len(tran)):
    search_bar = browser.find_element_by_id('BodyContent_txtSearchReferenceno')
    search_btn = browser.find_element_by_id('BodyContent_btnSearch')
    search_bar.send_keys(tran[i])
    search_btn.click()
    try:
        status = browser.find_element_by_xpath('//*[@id="BodyContent_gvDeposits"]/tbody/tr[2]/td[5]').text
    except:
        status = 'Nope'
    try:
        status1 = browser.find_element_by_xpath('//*[@id="BodyContent_gvDeposits"]/tbody/tr[3]/td[5]').text
        sh.cell(row = i+2, column = 6).value = status1
    except:
        pass
    try:
        status2 = browser.find_element_by_xpath('//*[@id="BodyContent_gvDeposits"]/tbody/tr[4]/td[5]').text
        sh.cell(row = i+2, column = 7).value = status2
    except:
        pass
    try:
        status3 = browser.find_element_by_xpath('//*[@id="BodyContent_gvDeposits"]/tbody/tr[5]/td[5]').text
        sh.cell(row = i+2, column = 8).value = status3
    except:
        pass
    try:
        status4 = browser.find_element_by_xpath('//*[@id="BodyContent_gvDeposits"]/tbody/tr[6]/td[5]').text
        sh.cell(row = i+2, column = 9).value = status4
    except:
        pass
    try:
        status5 = browser.find_element_by_xpath('//*[@id="BodyContent_gvDeposits"]/tbody/tr[7]/td[5]').text
        sh.cell(row = i+2, column = 10).value = status5
    except:
        pass
    try:
        status6 = browser.find_element_by_xpath('//*[@id="BodyContent_gvDeposits"]/tbody/tr[8]/td[5]').text
        sh.cell(row = i+2, column = 11).value = status6
    except:
        pass
    try:
        status7 = browser.find_element_by_xpath('//*[@id="BodyContent_gvDeposits"]/tbody/tr[9]/td[5]').text
        sh.cell(row = i+2, column = 12).value = status7
    except:
        pass
    try:
        status8 = browser.find_element_by_xpath('//*[@id="BodyContent_gvDeposits"]/tbody/tr[10]/td[5]').text
        sh.cell(row = i+2, column = 13).value = status8
    except:
        pass
    try:
        status9 = browser.find_element_by_xpath('//*[@id="BodyContent_gvDeposits"]/tbody/tr[11]/td[5]').text
        sh.cell(row = i+2, column = 14).value = status9
    except:
        pass
    try:
        status10 = browser.find_element_by_xpath('//*[@id="BodyContent_gvDeposits"]/tbody/tr[12]/td[5]').text
        sh.cell(row = i+2, column = 15).value = status10
    except:
        pass
    try:
        amt = browser.find_element_by_xpath('//*[@id="BodyContent_gvDeposits"]/tbody/tr[2]/td[4]').text
        amt = int(float(amt.replace(',','')))
    except:
        amt = 'NA'
    browser.find_element_by_id('BodyContent_btnSearchReset').click()
    stat.append(status)
    amount.append(amt)
    
for i in range(2,count+2):
    try:
        sh.cell(row=i, column=4).value = stat[i-2]
        sh.cell(row=i, column=5).value = amount[i-2]
        if sh.cell(row=i, column=4).value == 'Approved':
            sh.cell(row=i, column=4).fill = PatternFill(fill_type='solid', start_color='00ff00', end_color='00ff00')
        if sh.cell(row=i, column=4).value == 'Rejected':
            sh.cell(row=i, column=4).fill = PatternFill(fill_type='solid', start_color='ff0000', end_color='ff0000')
        if sh.cell(row=i, column=4).value == 'Nope':
            sh.cell(row = i, column = 4).value = 'No Entry Found'
            sh.cell(row=i, column=4).fill = PatternFill(fill_type='solid', start_color='ff0000', end_color='ffff00')
    except:
        pass
    wb.save(filename)
    

